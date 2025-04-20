import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from datetime import datetime

def criar_relatorio_excel(df, resultados, graficos, caminho_saida='output/relatorio_ofertas.xlsx'):
    """
    Cria um relatório Excel formatado com os dados processados.
    
    Args:
        df: DataFrame com os dados processados
        resultados: Dicionário com os resultados das análises
        graficos: Dicionário com os caminhos dos gráficos gerados
        caminho_saida: Caminho onde o arquivo Excel será salvo
    """
    print(f"Criando relatório Excel em {caminho_saida}...")
    
    # Criar diretório de saída se não existir
    os.makedirs(os.path.dirname(caminho_saida), exist_ok=True)
    
    # Criar um novo workbook
    wb = Workbook()
    
    # Configurar estilos
    titulo_estilo = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    subtitulo_estilo = Font(name='Arial', size=12, bold=True)
    cabecalho_estilo = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    texto_estilo = Font(name='Arial', size=10)
    
    titulo_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    cabecalho_fill = PatternFill(start_color='8EB4E3', end_color='8EB4E3', fill_type='solid')
    linha_alternada_fill = PatternFill(start_color='EBF1F8', end_color='EBF1F8', fill_type='solid')
    
    borda = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    centro_alinhamento = Alignment(horizontal='center', vertical='center')
    esquerda_alinhamento = Alignment(horizontal='left', vertical='center')
    direita_alinhamento = Alignment(horizontal='right', vertical='center')
    
    # Planilha de Resumo
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    
    # Título do relatório
    ws_resumo.merge_cells('A1:G1')
    ws_resumo['A1'] = "RELATÓRIO DE OFERTAS - " + datetime.now().strftime('%d/%m/%Y')
    ws_resumo['A1'].font = titulo_estilo
    ws_resumo['A1'].fill = titulo_fill
    ws_resumo['A1'].alignment = centro_alinhamento
    
    # Estatísticas gerais
    ws_resumo['A3'] = "Estatísticas Gerais"
    ws_resumo['A3'].font = subtitulo_estilo
    
    estatisticas = [
        ("Total de Produtos", resultados['total_produtos']),
        ("Preço Médio", f"R$ {resultados['preco_medio']:.2f}"),
        ("Desconto Médio", f"{resultados['desconto_medio']:.2f}%"),
        ("Economia Total", f"R$ {resultados['economia_total']:.2f}")
    ]
    
    for i, (label, valor) in enumerate(estatisticas, 4):
        ws_resumo[f'A{i}'] = label
        ws_resumo[f'A{i}'].font = texto_estilo
        ws_resumo[f'B{i}'] = valor
        ws_resumo[f'B{i}'].font = texto_estilo
        ws_resumo[f'B{i}'].alignment = direita_alinhamento
    
    # Adicionar gráficos
    ws_resumo['A9'] = "Análise Visual"
    ws_resumo['A9'].font = subtitulo_estilo
    
    # Adicionar imagem do gráfico de desconto por site
    if os.path.exists(graficos['desconto_por_site']):
        img = Image(graficos['desconto_por_site'])
        img.width = 500
        img.height = 300
        ws_resumo.add_image(img, 'A11')
    
    # Adicionar imagem do gráfico de distribuição por site
    if os.path.exists(graficos['distribuicao_por_site']):
        img = Image(graficos['distribuicao_por_site'])
        img.width = 500
        img.height = 300
        ws_resumo.add_image(img, 'A30')
    
    # Ajustar largura das colunas
    for col in range(1, 8):
        ws_resumo.column_dimensions[get_column_letter(col)].width = 15
    
    # Planilha de Melhores Ofertas
    ws_ofertas = wb.create_sheet(title="Melhores Ofertas")
    
    # Título
    ws_ofertas.merge_cells('A1:G1')
    ws_ofertas['A1'] = "MELHORES OFERTAS POR DESCONTO"
    ws_ofertas['A1'].font = titulo_estilo
    ws_ofertas['A1'].fill = titulo_fill
    ws_ofertas['A1'].alignment = centro_alinhamento
    
    # Cabeçalhos
    cabecalhos = ["Site", "Título", "Preço Atual", "Preço Antigo", "Desconto", "Economia", "Link"]
    for col, cabecalho in enumerate(cabecalhos, 1):
        ws_ofertas.cell(row=3, column=col, value=cabecalho)
        cell = ws_ofertas.cell(row=3, column=col)
        cell.font = cabecalho_estilo
        cell.fill = cabecalho_fill
        cell.alignment = centro_alinhamento
        cell.border = borda
    
    # Dados das melhores ofertas
    melhores_ofertas = resultados['melhores_ofertas']
    for i, (_, oferta) in enumerate(melhores_ofertas.iterrows(), 4):
        ws_ofertas.cell(row=i, column=1, value=oferta['site'])
        ws_ofertas.cell(row=i, column=2, value=oferta['titulo'])
        ws_ofertas.cell(row=i, column=3, value=f"R$ {oferta['preco_atual']:.2f}")
        ws_ofertas.cell(row=i, column=4, value=f"R$ {oferta['preco_antigo']:.2f}" if pd.notna(oferta['preco_antigo']) else "N/A")
        ws_ofertas.cell(row=i, column=5, value=f"{oferta['desconto']:.2f}%")
        ws_ofertas.cell(row=i, column=6, value=f"R$ {oferta['economia']:.2f}" if pd.notna(oferta['economia']) else "N/A")
        ws_ofertas.cell(row=i, column=7, value=oferta['link'])
        
        # Aplicar estilos
        for col in range(1, 8):
            cell = ws_ofertas.cell(row=i, column=col)
            cell.font = texto_estilo
            cell.border = borda
            
            # Alinhamento específico por coluna
            if col in [1, 2, 7]:  # Site, Título, Link
                cell.alignment = esquerda_alinhamento
            elif col in [3, 4, 5, 6]:  # Valores numéricos
                cell.alignment = direita_alinhamento
            
            # Linhas alternadas
            if i % 2 == 0:
                cell.fill = linha_alternada_fill
    
    # Ajustar largura das colunas
    ws_ofertas.column_dimensions['A'].width = 15  # Site
    ws_ofertas.column_dimensions['B'].width = 40  # Título
    ws_ofertas.column_dimensions['C'].width = 15  # Preço Atual
    ws_ofertas.column_dimensions['D'].width = 15  # Preço Antigo
    ws_ofertas.column_dimensions['E'].width = 15  # Desconto
    ws_ofertas.column_dimensions['F'].width = 15  # Economia
    ws_ofertas.column_dimensions['G'].width = 40  # Link
    
    # Planilha de Análise por Site
    ws_sites = wb.create_sheet(title="Análise por Site")
    
    # Título
    ws_sites.merge_cells('A1:E1')
    ws_sites['A1'] = "ANÁLISE POR SITE"
    ws_sites['A1'].font = titulo_estilo
    ws_sites['A1'].fill = titulo_fill
    ws_sites['A1'].alignment = centro_alinhamento
    
    # Cabeçalhos
    cabecalhos = ["Site", "Quantidade", "Preço Médio", "Desconto Médio", "Economia Total"]
    for col, cabecalho in enumerate(cabecalhos, 1):
        ws_sites.cell(row=3, column=col, value=cabecalho)
        cell = ws_sites.cell(row=3, column=col)
        cell.font = cabecalho_estilo
        cell.fill = cabecalho_fill
        cell.alignment = centro_alinhamento
        cell.border = borda
    
    # Dados da análise por site
    analise_por_site = resultados['analise_por_site'].reset_index()
    for i, (_, site_data) in enumerate(analise_por_site.iterrows(), 4):
        ws_sites.cell(row=i, column=1, value=site_data['site'])
        ws_sites.cell(row=i, column=2, value=site_data['quantidade'])
        ws_sites.cell(row=i, column=3, value=f"R$ {site_data['preco_atual']:.2f}")
        ws_sites.cell(row=i, column=4, value=f"{site_data['desconto']:.2f}%")
        ws_sites.cell(row=i, column=5, value=f"R$ {site_data['economia']:.2f}")
        
        # Aplicar estilos
        for col in range(1, 6):
            cell = ws_sites.cell(row=i, column=col)
            cell.font = texto_estilo
            cell.border = borda
            
            # Alinhamento específico por coluna
            if col == 1:  # Site
                cell.alignment = esquerda_alinhamento
            else:  # Valores numéricos
                cell.alignment = direita_alinhamento
            
            # Linhas alternadas
            if i % 2 == 0:
                cell.fill = linha_alternada_fill
    
    # Adicionar gráfico de barras para desconto médio por site
    chart = BarChart()
    chart.title = "Desconto Médio por Site"
    chart.y_axis.title = "Desconto (%)"
    chart.x_axis.title = "Site"
    
    data = Reference(ws_sites, min_col=4, min_row=3, max_row=3+len(analise_por_site), max_col=4)
    cats = Reference(ws_sites, min_col=1, min_row=4, max_row=3+len(analise_por_site))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 15
    
    ws_sites.add_chart(chart, "G3")
    
    # Ajustar largura das colunas
    ws_sites.column_dimensions['A'].width = 15  # Site
    ws_sites.column_dimensions['B'].width = 15  # Quantidade
    ws_sites.column_dimensions['C'].width = 15  # Preço Médio
    ws_sites.column_dimensions['D'].width = 15  # Desconto Médio
    ws_sites.column_dimensions['E'].width = 15  # Economia Total
    
    # Planilha com todos os produtos
    ws_todos = wb.create_sheet(title="Todos os Produtos")
    
    # Título
    ws_todos.merge_cells('A1:H1')
    ws_todos['A1'] = "LISTA COMPLETA DE PRODUTOS"
    ws_todos['A1'].font = titulo_estilo
    ws_todos['A1'].fill = titulo_fill
    ws_todos['A1'].alignment = centro_alinhamento
    
    # Cabeçalhos
    cabecalhos = ["Site", "Título", "Categoria", "Preço Atual", "Preço Antigo", "Desconto", "Economia", "Link"]
    for col, cabecalho in enumerate(cabecalhos, 1):
        ws_todos.cell(row=3, column=col, value=cabecalho)
        cell = ws_todos.cell(row=3, column=col)
        cell.font = cabecalho_estilo
        cell.fill = cabecalho_fill
        cell.alignment = centro_alinhamento
        cell.border = borda
    
    # Dados de todos os produtos
    for i, (_, produto) in enumerate(df.iterrows(), 4):
        ws_todos.cell(row=i, column=1, value=produto['site'])
        ws_todos.cell(row=i, column=2, value=produto['titulo'])
        ws_todos.cell(row=i, column=3, value=produto['categoria'])
        ws_todos.cell(row=i, column=4, value=f"R$ {produto['preco_atual']:.2f}")
        ws_todos.cell(row=i, column=5, value=f"R$ {produto['preco_antigo']:.2f}" if pd.notna(produto['preco_antigo']) else "N/A")
        ws_todos.cell(row=i, column=6, value=f"{produto['desconto']:.2f}%" if pd.notna(produto['desconto']) else "N/A")
        ws_todos.cell(row=i, column=7, value=f"R$ {produto['economia']:.2f}" if pd.notna(produto['economia']) else "N/A")
        ws_todos.cell(row=i, column=8, value=produto['link'])
        
        # Aplicar estilos
        for col in range(1, 9):
            cell = ws_todos.cell(row=i, column=col)
            cell.font = texto_estilo
            cell.border = borda
            
            # Alinhamento específico por coluna
            if col in [1, 2, 3, 8]:  # Site, Título, Categoria, Link
                cell.alignment = esquerda_alinhamento
            elif col in [4, 5, 6, 7]:  # Valores numéricos
                cell.alignment = direita_alinhamento
            
            # Linhas alternadas
            if i % 2 == 0:
                cell.fill = linha_alternada_fill
    
    # Ajustar largura das colunas
    ws_todos.column_dimensions['A'].width = 15  # Site
    ws_todos.column_dimensions['B'].width = 40  # Título
    ws_todos.column_dimensions['C'].width = 15  # Categoria
    ws_todos.column_dimensions['D'].width = 15  # Preço Atual
    ws_todos.column_dimensions['E'].width = 15  # Preço Antigo
    ws_todos.column_dimensions['F'].width = 15  # Desconto
    ws_todos.column_dimensions['G'].width = 15  # Economia
    ws_todos.column_dimensions['H'].width = 40  # Link
    
    # Salvar o workbook
    wb.save(caminho_saida)
    
    print(f"Relatório Excel criado com sucesso em {caminho_saida}")
    
    return caminho_saida

def gerar_relatorio_excel(df=None, resultados=None, graficos=None):
    """Função principal para gerar o relatório Excel."""
    # Se não foram fornecidos dados, carregar do arquivo
    if df is None or resultados is None or graficos is None:
        import processador
        df, resultados, graficos = processador.processar_dados()
    
    # Criar relatório Excel
    caminho_relatorio = criar_relatorio_excel(df, resultados, graficos)
    
    return caminho_relatorio

if __name__ == "__main__":
    gerar_relatorio_excel()

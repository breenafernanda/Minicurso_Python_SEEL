#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Exemplo Prático: Criação de relatório formatado em Excel com OpenPyXL

Este script demonstra como usar a biblioteca OpenPyXL para criar relatórios
Excel profissionais e bem formatados a partir de dados de produtos.

Conceitos demonstrados:
- Criação e manipulação de workbooks e planilhas
- Formatação avançada de células (fontes, cores, bordas)
- Mesclagem de células para títulos e cabeçalhos
- Criação de tabelas formatadas
- Adição de fórmulas e funções
- Criação de gráficos
- Adição de imagens e logos
- Proteção de planilhas

Autor: Equipe do Minicurso de Automação com Python
Data: Abril de 2025
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Protection
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.image import Image
import json
import os
import random
from datetime import datetime

# Diretórios
DADOS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'dados')
IMAGENS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'imagens')
os.makedirs(DADOS_DIR, exist_ok=True)
os.makedirs(IMAGENS_DIR, exist_ok=True)

# Cores para uso no relatório
CORES = {
    'azul_claro': 'B3E0FF',
    'azul_escuro': '0078D7',
    'verde_claro': 'C6EFCE',
    'verde_escuro': '006100',
    'vermelho_claro': 'FFC7CE',
    'vermelho_escuro': '9C0006',
    'amarelo_claro': 'FFEB9C',
    'amarelo_escuro': '9C6500',
    'cinza_claro': 'F2F2F2',
    'cinza_escuro': '595959',
    'branco': 'FFFFFF',
    'preto': '000000'
}

def carregar_dados(arquivo=None):
    """
    Carrega dados de produtos de um arquivo JSON.
    
    Args:
        arquivo (str, optional): Nome do arquivo específico a ser carregado.
            Se None, tenta carregar qualquer arquivo de produtos disponível.
    
    Returns:
        list: Lista de dicionários com os dados dos produtos
    """
    # Se nenhum arquivo específico foi fornecido, procurar por arquivos de produtos
    if arquivo is None:
        arquivos_json = [f for f in os.listdir(DADOS_DIR) if f.startswith('produtos_') and f.endswith('.json')]
        if not arquivos_json:
            print("Nenhum arquivo de produtos encontrado. Gerando dados de exemplo...")
            return gerar_dados_exemplo()
        
        # Usar o arquivo mais recente
        arquivo = max(arquivos_json, key=lambda f: os.path.getmtime(os.path.join(DADOS_DIR, f)))
    
    caminho_arquivo = os.path.join(DADOS_DIR, arquivo)
    print(f"Carregando dados de: {caminho_arquivo}")
    
    try:
        # Carregar dados do JSON
        with open(caminho_arquivo, 'r', encoding='utf-8') as f:
            dados = json.load(f)
        
        print(f"Dados carregados com sucesso: {len(dados)} produtos")
        return dados
    
    except FileNotFoundError:
        print(f"Arquivo não encontrado: {caminho_arquivo}")
        return gerar_dados_exemplo()
    
    except Exception as e:
        print(f"Erro ao carregar dados: {str(e)}")
        return gerar_dados_exemplo()

def gerar_dados_exemplo():
    """
    Gera dados fictícios de produtos para demonstração.
    
    Returns:
        list: Lista de dicionários com dados fictícios
    """
    print("Gerando dados de exemplo...")
    
    # Criar dados fictícios
    dados = []
    categorias = ['Informática', 'Eletrônicos', 'Celulares', 'Acessórios', 'Casa']
    
    for i in range(50):
        categoria = random.choice(categorias)
        preco_original = random.uniform(100, 5000)
        desconto = random.uniform(0, 0.4)
        preco_atual = preco_original * (1 - desconto)
        
        produto = {
            'nome': f'Produto Exemplo {i+1} - {categoria}',
            'categoria': categoria,
            'preco_atual': round(preco_atual, 2),
            'preco_original': round(preco_original, 2),
            'desconto': round(desconto * 100, 2),
            'avaliacao': round(random.uniform(1, 5), 1),
            'num_avaliacoes': random.randint(0, 1000),
            'link': f'https://exemplo.com/produto{i+1}',
            'imagem': f'https://exemplo.com/imagens/produto{i+1}.jpg',
            'disponivel': random.choice([True, False]),
            'vendedor': f'Vendedor {random.randint(1, 10)}',
            'data_extracao': datetime.now().strftime('%Y-%m-%d')
        }
        dados.append(produto)
    
    # Salvar dados de exemplo
    caminho_arquivo = os.path.join(DADOS_DIR, 'produtos_exemplo.json')
    with open(caminho_arquivo, 'w', encoding='utf-8') as f:
        json.dump(dados, f, ensure_ascii=False, indent=4)
    print(f"Dados de exemplo salvos em: {caminho_arquivo}")
    
    return dados

def criar_relatorio_excel(dados, nome_arquivo=None):
    """
    Cria um relatório Excel formatado com os dados dos produtos.
    
    Args:
        dados (list): Lista de dicionários com os dados dos produtos
        nome_arquivo (str, optional): Nome do arquivo Excel a ser criado.
            Se None, gera um nome com timestamp.
    
    Returns:
        str: Caminho do arquivo Excel criado
    """
    if not nome_arquivo:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f"relatorio_produtos_{timestamp}.xlsx"
    
    caminho_arquivo = os.path.join(DADOS_DIR, nome_arquivo)
    
    print(f"Criando relatório Excel: {caminho_arquivo}")
    
    # Criar um novo workbook
    wb = Workbook()
    
    # Configurar a planilha principal
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    
    # Criar planilhas adicionais
    ws_produtos = wb.create_sheet(title="Produtos")
    ws_categorias = wb.create_sheet(title="Análise por Categoria")
    ws_graficos = wb.create_sheet(title="Gráficos")
    
    # Formatar e preencher cada planilha
    formatar_planilha_resumo(ws_resumo, dados)
    formatar_planilha_produtos(ws_produtos, dados)
    formatar_planilha_categorias(ws_categorias, dados)
    formatar_planilha_graficos(ws_graficos, dados)
    
    # Proteger as planilhas (somente leitura)
    for ws in wb.worksheets:
        ws.protection.sheet = True
        ws.protection.password = 'python'  # Senha simples para demonstração
    
    # Salvar o workbook
    wb.save(caminho_arquivo)
    print(f"Relatório Excel criado com sucesso: {caminho_arquivo}")
    
    return caminho_arquivo

def formatar_planilha_resumo(ws, dados):
    """
    Formata e preenche a planilha de resumo.
    
    Args:
        ws (Worksheet): Objeto da planilha
        dados (list): Lista de dicionários com os dados dos produtos
    """
    print("Formatando planilha de Resumo...")
    
    # Definir larguras das colunas
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    
    # Título do relatório
    ws.merge_cells('A1:D2')
    cell = ws['A1']
    cell.value = "RELATÓRIO DE PRODUTOS"
    cell.font = Font(name='Arial', size=16, bold=True, color=CORES['branco'])
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = PatternFill(start_color=CORES['azul_escuro'], end_color=CORES['azul_escuro'], fill_type='solid')
    
    # Data do relatório
    ws.merge_cells('A3:D3')
    cell = ws['A3']
    cell.value = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    cell.font = Font(name='Arial', size=10, italic=True)
    cell.alignment = Alignment(horizontal='center')
    
    # Adicionar espaço
    ws.merge_cells('A4:D4')
    
    # Estatísticas gerais - Título
    ws.merge_cells('A5:D5')
    cell = ws['A5']
    cell.value = "ESTATÍSTICAS GERAIS"
    cell.font = Font(name='Arial', size=12, bold=True)
    cell.alignment = Alignment(horizontal='center')
    cell.fill = PatternFill(start_color=CORES['cinza_claro'], end_color=CORES['cinza_claro'], fill_type='solid')
    
    # Borda para o título
    thin_border = Border(
        left=Side(style='thin', color=CORES['cinza_escuro']),
        right=Side(style='thin', color=CORES['cinza_escuro']),
        top=Side(style='thin', color=CORES['cinza_escuro']),
        bottom=Side(style='thin', color=CORES['cinza_escuro'])
    )
    cell.border = thin_border
    
    # Estatísticas gerais - Dados
    metricas = [
        ("Total de produtos", len(dados)),
        ("Preço médio", f"R$ {sum(p['preco_atual'] for p in dados) / len(dados):.2f}"),
        ("Preço mínimo", f"R$ {min(p['preco_atual'] for p in dados):.2f}"),
        ("Preço máximo", f"R$ {max(p['preco_atual'] for p in dados):.2f}"),
        ("Desconto médio", f"{sum(p.get('desconto', 0) for p in dados) / len(dados):.2f}%"),
        ("Maior desconto", f"{max(p.get('desconto', 0) for p in dados):.2f}%"),
        ("Produtos com desconto", f"{sum(1 for p in dados if p.get('desconto', 0) > 0)} ({sum(1 for p in dados if p.get('desconto', 0) > 0) / len(dados) * 100:.1f}%)"),
        ("Avaliação média", f"{sum(p.get('avaliacao', 0) for p in dados if 'avaliacao' in p) / sum(1 for p in dados if 'avaliacao' in p):.1f}/5.0" if any('avaliacao' in p for p in dados) else "N/A")
    ]
    
    for i, (metrica, valor) in enumerate(metricas, 6):
        # Métrica
        cell = ws[f'A{i}']
        cell.value = metrica
        cell.font = Font(name='Arial', size=11)
        cell.alignment = Alignment(horizontal='left')
        
        # Valor
        cell = ws[f'B{i}']
        cell.value = valor
        cell.font = Font(name='Arial', size=11, bold=True)
        cell.alignment = Alignment(horizontal='left')
    
    # Adicionar espaço
    ws.merge_cells(f'A{6+len(metricas)}:D{6+len(metricas)}')
    
    # Top 5 produtos mais caros - Título
    row = 7 + len(metricas)
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "TOP 5 PRODUTOS MAIS CAROS"
    cell.font = Font(name='Arial', size=12, bold=True)
    cell.alignment = Alignment(horizontal='center')
    cell.fill = PatternFill(start_color=CORES['cinza_claro'], end_color=CORES['cinza_claro'], fill_type='solid')
    cell.border = thin_border
    
    # Top 5 produtos mais caros - Cabeçalho
    row += 1
    headers = ["Nome do Produto", "Preço", "Desconto", "Categoria"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(name='Arial', size=11, bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color=CORES['azul_claro'], end_color=CORES['azul_claro'], fill_type='solid')
        cell.border = thin_border
    
    # Top 5 produtos mais caros - Dados
    produtos_ordenados = sorted(dados, key=lambda p: p['preco_atual'], reverse=True)[:5]
    for i, produto in enumerate(produtos_ordenados, 1):
        row += 1
        
        # Nome
        cell = ws.cell(row=row, column=1)
        cell.value = produto['nome']
        cell.font = Font(name='Arial', size=10)
        cell.alignment = Alignment(horizontal='left')
        cell.border = thin_border
        
        # Preço
        cell = ws.cell(row=row, column=2)
        cell.value = f"R$ {produto['preco_atual']:.2f}"
        cell.font = Font(name='Arial', size=10, bold=True)
        cell.alignment = Alignment(horizontal='right')
        cell.border = thin_border
        
        # Desconto
        cell = ws.cell(row=row, column=3)
        cell.value = f"{produto.get('desconto', 0):.2f}%"
        cell.font = Font(name='Arial', size=10)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        
        # Categoria
        cell = ws.cell(row=row, column=4)
        cell.value = produto.get('categoria', 'N/A')
        cell.font = Font(name='Arial', size=10)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Adicionar espaço
    row += 2
    ws.merge_cells(f'A{row}:D{row}')
    
    # Top 5 produtos com maior desconto - Título
    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "TOP 5 PRODUTOS COM MAIOR DESCONTO"
    cell.font = Font(name='Arial', size=12, bold=True)
    cell.alignment = Alignment(horizontal='center')
    cell.fill = PatternFill(start_color=CORES['cinza_claro'], end_color=CORES['cinza_claro'], fill_type='solid')
    cell.border = thin_border
    
    # Top 5 produtos com maior desconto - Cabeçalho
    row += 1
    headers = ["Nome do Produto", "Preço Original", "Preço Atual", "Desconto"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(name='Arial', size=11, bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color=CORES['verde_claro'], end_color=CORES['verde_claro'], fill_type='solid')
        cell.border = thin_border
    
    # Top 5 produtos com maior desconto - Dados
    produtos_com_desconto = [p for p in dados if p.get('desconto', 0) > 0]
    produtos_ordenados = sorted(produtos_com_desconto, key=lambda p: p.get('desconto', 0), reverse=True)[:5]
    for i, produto in enumerate(produtos_ordenados, 1):
        row += 1
        
        # Nome
        cell = ws.cell(row=row, column=1)
        cell.value = produto['nome']
        cell.font = Font(name='Arial', size=10)
        cell.alignment = Alignment(horizontal='left')
        cell.border = thin_border
        
        # Preço Original
        cell = ws.cell(row=row, column=2)
        cell.value = f"R$ {produto.get('preco_original', produto['preco_atual']):.2f}"
        cell.font = Font(name='Arial', size=10)
        cell.alignment = Alignment(horizontal='right')
        cell.border = thin_border
        
        # Preço Atual
        cell = ws.cell(row=row, column=3)
        cell.value = f"R$ {produto['preco_atual']:.2f}"
        cell.font = Font(name='Arial', size=10, bold=True)
        cell.alignment = Alignment(horizontal='right')
        cell.border = thin_border
        
        # Desconto
        cell = ws.cell(row=row, column=4)
        cell.value = f"{produto.get('desconto', 0):.2f}%"
        cell.font = Font(name='Arial', size=10, bold=True, color=CORES['verde_escuro'])
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color=CORES['verde_claro'], end_color=CORES['verde_claro'], fill_type='solid')
        cell.border = thin_border
    
    # Adicionar nota de rodapé
    row += 3
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "Nota: Este relatório foi gerado automaticamente usando Python e OpenPyXL."
    cell.font = Font(name='Arial', size=9, italic=True)
    cell.alignment = Alignment(horizontal='center')

def formatar_planilha_produtos(ws, dados):
    """
    Formata e preenche a planilha de produtos.
    
    Args:
        ws (Worksheet): Objeto da planilha
        dados (list): Lista de dicionários com os dados dos produtos
    """
    print("Formatando planilha de Produtos...")
    
    # Definir larguras das colunas
    colunas = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
    larguras = [5, 40, 15, 15, 12, 12, 15, 15, 30, 20]
    for col, largura in zip(colunas, larguras):
        ws.column_dimensions[col].width = largura
    
    # Título da planilha
    ws.merge_cells('A1:J1')
    cell = ws['A1']
    cell.value = "LISTA COMPLETA DE PRODUTOS"
    cell.font = Font(name='Arial', size=14, bold=True, color=CORES['branco'])
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = PatternFill(start_color=CORES['azul_escuro'], end_color=CORES['azul_escuro'], fill_type='solid')
    
    # Cabeçalho da tabela
    headers = ["#", "Nome do Produto", "Preço Atual", "Preço Original", "Desconto", "Avaliação", "Categoria", "Vendedor", "Link", "Data Extração"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column
(Content truncated due to size limit. Use line ranges to read in chunks)
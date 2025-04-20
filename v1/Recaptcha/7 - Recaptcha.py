"""INSTALAR BIBLIOTECAS NECESSARIAS
pip install selenium
pip install ChromeDriverManager
"""
# importar bibliotecas necessárias
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import time
from colorama import Fore, Style
import openpyxl
from openpyxl import Workbook
import pandas as pd
from fpdf import FPDF
import requests
from io import BytesIO
from PIL import Image

# Função para imprimir colorido no terminal 
def imprimir(texto, valor):
    print(f"{Fore.LIGHTMAGENTA_EX}{texto}: {Fore.CYAN}{valor}{Style.RESET_ALL}")


""" Script do Projeto - Selenium """
# Inicializa WebDriver Chrome
def abrir_navegador():
    driver = webdriver.Chrome(Options())
    driver.maximize_window()
    return driver

# SCRIPT AUXLIAR 1ª etapa - Extrair dados de 1 produto - retorna {produto}
def extrair_dados_selenium(driver, url):
    try:
        # acessar url 
        driver.get(url)
        time.sleep(5)

        """IDENTIFICAR SELETORES DE ELEMENTO WEB"""
        id_nome_produto = 'productTitle'
        selector_porcentagem_desconto = '#corePriceDisplay_desktop_feature_div > div.a-section.a-spacing-none.aok-align-center.aok-relative > span.a-size-large.a-color-price.savingPriceOverride.aok-align-center.reinventPriceSavingsPercentageMargin.savingsPercentage'
        selector_preço_desconto = '#corePriceDisplay_desktop_feature_div > div.a-section.a-spacing-none.aok-align-center.aok-relative > span.a-price.aok-align-center.reinventPricePriceToPayMargin.priceToPay > span:nth-child(2)'
        selector_preço_original = '#corePriceDisplay_desktop_feature_div > div.a-section.a-spacing-small.aok-align-center > span > span.aok-relative > span.a-size-small.a-color-secondary.aok-align-center.basisPrice > span > span:nth-child(2)'
        xpath_imagem_produto = '//*[@id="landingImage"]'
        
        # extrair nome do produto
        nome_do_produto = driver.find_element(By.ID, id_nome_produto).text
        imprimir("Nome do Produto", nome_do_produto)

        # Porcentagem de desconto 
        try:
            porcentagem_desconto = driver.find_element(By.CSS_SELECTOR, selector_porcentagem_desconto).text
            imprimir("Desconto (%)", porcentagem_desconto)
        except:
            porcentagem_desconto = '-0%'
            imprimir("Desconto (%)", porcentagem_desconto)

        # extrair preço original
        try:
            preço_original = driver.find_element(By.CSS_SELECTOR, selector_preço_original).text
            imprimir("Preço Original", preço_original)
        except: 
            preço_original = driver.find_element(By.XPATH, '//*[@id="corePrice_desktop"]/div/table/tbody/tr/td[2]').text
            imprimir("Preço Original", preço_original)


        # extrair preço com desconto
        try:
            preço_desconto = driver.find_element(By.CSS_SELECTOR, selector_preço_desconto).text.replace("\n", ',')
            imprimir("Preço com Desconto", preço_desconto)
        except: 
            preço_desconto = preço_original
        
        # extrair imagem 
        elemento_imagem = driver.find_element(By.XPATH, xpath_imagem_produto).get_attribute('src')
        imprimir("Imagem (scr)", elemento_imagem)

        # Criar dicionário com todos os dados do Produto
        produto = {
            "Nome": nome_do_produto,
            "Desconto %": porcentagem_desconto,
            "Preço Original": preço_original,
            "Preço com Desconto": preço_desconto,
            "URL da Imagem": elemento_imagem
        }
        print(f'{Fore.LIGHTMAGENTA_EX}------------------------------------{Style.RESET_ALL}')
        return produto
    
    except: 
        img_recaptcha = 'body > div > div.a-row.a-spacing-double-large > div.a-section > div > div > form > div.a-row.a-spacing-large > div > div > div.a-row.a-text-center > img'
        try: 
            while True:
                try:
                    recaptcha = driver.find_element(By.CSS_SELECTOR, img_recaptcha).get_attribute('src')
                    print(f'{Fore.RED}Resolva o Recaptcha{Style.RESET_ALL}')
                    resolver_recaptcha(recaptcha, driver)
                    time.sleep(30)
                except: break 
            return extrair_dados_selenium(driver, url)


        except: print('Recaptcha não localizado.')
        time.sleep(1000)

# SCRIPT PRINCIPAL 1ª etapa - Extrair dados de todos produtos - retorna [produtos]
def extrair_produtos():
    # Inicializar webdriver
    driver = abrir_navegador()
    
    # Definir URL's
    url1 = 'https://www.amazon.com.br/Recado-3M-Autoadesivo-Cole%C3%A7%C3%A3o-Supernova/dp/B09YYGCPGV?ref=dlx_deals_dg_dcl_B09YYGCPGV_dt_sl14_ac&pf_rd_r=ZRWD59DBNHXQHY5Q6966&pf_rd_p=2d04dc0d-d8ca-40f6-b7ec-b3fce0874cac'
    url2 = 'https://www.amazon.com.br/Xixaomiro-Teclado-Polegada-1920x1200-8000mAh/dp/B0D1FN49HJ?ref=dlx_deals_dg_dcl_B0D1FN49HJ_dt_sl14_ac&pf_rd_r=DVX7TFBWJ55KHMK0NE8A&pf_rd_p=2d04dc0d-d8ca-40f6-b7ec-b3fce0874cac'
    url3 = 'https://www.amazon.com.br/GameSir-G7-SE-Controle-Joysticks/dp/B0C7GW9F88?ref=dlx_deals_dg_dcl_B0C7GW9F88_dt_sl14_ac&pf_rd_r=M50A1FM4X6NTCE7ASXW4&pf_rd_p=2d04dc0d-d8ca-40f6-b7ec-b3fce0874cac'
    url4 = 'https://www.amazon.com.br/Bicicleta-Caloi-Evora-CALOI-Vermelho/dp/B0CYQC2ZB8?ref=dlx_deals_dg_dcl_B0CYQC2ZB8_dt_sl14_ac&pf_rd_r=B8WY63P3MVJH5W4QG2NG&pf_rd_p=2d04dc0d-d8ca-40f6-b7ec-b3fce0874cac'
    url5 = 'https://www.amazon.com.br/Caderno-Universit%C3%A1rio-Tilibra-305421-Mat%C3%A9rias/dp/B07VFWFS4P?ref=dlx_deals_dg_dcl_B07VFWFS4P_dt_sl14_ac&pf_rd_r=DZPGKCTDC6BSV3CKNYT6&pf_rd_p=2d04dc0d-d8ca-40f6-b7ec-b3fce0874cac'
    url6 = 'https://www.amazon.com.br/Microfone-Din%C3%A2mico-Fifine-para-Jogos/dp/B0C74G9RLH?ref=dlx_deals_dg_dcl_B0C74G9RLH_dt_sl14_ac&pf_rd_r=B8WY63P3MVJH5W4QG2NG&pf_rd_p=2d04dc0d-d8ca-40f6-b7ec-b3fce0874cac'
    # vetor_de_urls = [url1, url2, url3, url4, url5, url6]
    vetor_de_urls = extrair_urls_de_produtos(driver)

    # Iniciar lista de
    produtos = []

    for url in vetor_de_urls:
        produto = extrair_dados_selenium(driver, url)
        produtos.append(produto)

    # Fechar webdriver
    driver.quit()
    
    # Retorna o vetor de produtos
    return produtos

def extrair_urls_de_produtos(driver):
    try:
        url_ofertas_do_dia = 'https://www.amazon.com.br/deals?ref_=nav_cs_gb'
        driver.get(url_ofertas_do_dia)
        time.sleep(5)
        urls = []
        produtos_grid = driver.find_elements(By.CSS_SELECTOR, 'a[data-testid="product-card-link"]')
        for produto in produtos_grid:
            url = produto.get_attribute('href')
            text = produto.text
            print(f'{url}\n{text}\n\n')
            urls.append(url)
        print(urls)
        return urls
    except: return extrair_urls_de_produtos(driver)

def resolver_recaptcha(scr, driver):
    # baixar recaptcha do scr recebido via request
    # resolver o recaptcha e encontrar captcha=resolvido
    captcha = 'PASDKAP'
    input_captcha = driver.find_element(By.ID, 'captchacharacters')
    input_captcha.send_keys(captcha)

    botao_continuar = driver.find_element(By.CSS_SELECTOR, 'body > div > div.a-row.a-spacing-double-large > div.a-section > div > div > form > div.a-section.a-spacing-extra-large > div > span > span > button')
    botao_continuar.click()
    print(f'Tentando solucionar captcha {captcha}...')

""" Script do Projeto - Openpyxl """
# Salvar [produtos] em "planilha.xlsx"
def salvar_planilha(produtos, nome_arquivo='planilha.xlsx'):
    # Cria um novo Workbook
    wb = Workbook()
    # Seleciona a planilha ativa
    ws = wb.active
    # Define o título da planilha
    ws.title = 'Produtos'

    # Adiciona o cabeçalho
    cabecalho = ['Nome', 'Desconto %', 'Preço Original', 'Preço com Desconto', 'URL da Imagem']  
    ws.append(cabecalho)

    # Adiciona os dados dos produtos
    for produto in produtos:
        ws.append([produto['Nome'], produto['Desconto %'], produto['Preço Original'], produto['Preço com Desconto'], produto['URL da Imagem']])
        # Ajuste as chaves conforme a estrutura do dicionário 'produto'

    # Salva o arquivo Excel
    wb.save(nome_arquivo)
    imprimir(f'Planilha salva como', f'{nome_arquivo}')
    return nome_arquivo


""" Script do Projeto - Pandas """
def obter_lista_produtos(nome_arquivo='planilha.xlsx'):
    # Lê o arquivo Excel utilizando pandas
    try:
        df = pd.read_excel(nome_arquivo)
        
        # Converte o DataFrame em uma lista de dicionários
        produtos = df.to_dict(orient='records')
        
        # Imprime os produtos para validação
        # imprimir("Produtos obtidos do Excel", produtos)
        
        return produtos
    except FileNotFoundError:
        print(f"Erro: O arquivo '{nome_arquivo}' não foi encontrado.")
        return []


""" Script do Projeto - PDF """
# Classe personalizada para criar o PDF
class PDF(FPDF):
    def header(self):
        self.set_font('Arial', 'B', 14)
        self.cell(0, 10, 'OFERTAS DO DIA', 0, 1, 'C')
        self.ln(10)

    def add_product(self, produto):
        self.set_font('Arial', '', 15)
        
        # Adicionar Imagem
        if produto['URL da Imagem']:
            try:
                response = requests.get(produto['URL da Imagem'])
                # path = produto['URL da Imagem'].split('_')[-2]
                image_path = produto['URL da Imagem'].split('/')[-1]  # Pega a última parte da URL

                if response.status_code == 200:
                    image = Image.open(BytesIO(response.content))
                    image.save(image_path)
                    # Dimensões da imagem no PDF
                    img_width = 50  # Largura desejada da imagem no PDF
                    largura_pagina = self.w  # Largura total da página
                    img_x = (largura_pagina - img_width) / 2  # Centralizar horizontalmente

                    # Adicionar a imagem
                    self.image(image_path, x=img_x, w=img_width)
                    imprimir('Imagem inserida', img_x)

            except:
                self.cell(0, 10, "Imagem não disponível.", ln=1)

        # Nome do Produto com quebra de linha (nomes muito grandes)
        self.set_font('Arial', 'B', 12)
        self.multi_cell(0, 10, f"{produto['Nome']}", align='C')
        self.set_font('Arial', '', 12)
        # Porcentagem de Desconto
        self.cell(0, 10, f"{produto['Desconto %']} DE DESCONTO-",align='C', ln=1)
        # Preço Original e Preço com Desconto
        self.cell(0, 10, f"Preço Original: {produto['Preço Original']}" + "    " + f"Preço com Desconto: {produto['Preço com Desconto']}", align='C',ln=1)
        self.cell(0, 10, f"-----------------------", ln=2)
        
        self.ln(50)  # Espaçamento entre produtos

# Função para criar o PDF
def criar_pdf(produtos, nome_arquivo='ofertas.pdf'):
    pdf = PDF()
    pdf.add_page()

    # Adicionar produtos no PDF
    for produto in produtos:
        pdf.add_product(produto)
    
    # Salvar o PDF
    pdf.output(nome_arquivo)
    imprimir(f"PDF salvo como", f"{nome_arquivo}")

# Script Execução
if __name__ == "__main__": 
    # 1ª etapa (Extrair dados da Web)
    produtos = extrair_produtos()

    # 2ª etapa (Salvar dados no excel)
    planilha = salvar_planilha(produtos)

    # 3ª etapa (Ler dados do excel)
    produtos_excel = obter_lista_produtos()

    # 4ª etapa (Salvar dados num PDF)
    arquivo_pdf = criar_pdf(produtos_excel)
    
